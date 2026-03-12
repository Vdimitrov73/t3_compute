"""
CDS T3 PDF Parser
Extracts distribution data from CDS T3 tax PDFs and saves as JSON.

Usage:
    pip install pdfplumber openpyxl
    python parse_t3_pdfs.py [--config config.json] [--year 2025] [--funds VBAL ZCN]

Output: one JSON file per fund in <base_dir>/<year>/distributions/
        one Excel ACB helper file per fund in the same folder

Calculation modes:
    RATE    — values are dollar amounts per unit (VBAL, VRE, CPD)
    PERCENT — values are % of total, converted to dollars (ZCN)

Config files (all in same directory as this script):
    config.json        — tax year, base_dir, fund list, column indices
    funds.json         — per-fund metadata (calc_method_override)
"""

import argparse
import pdfplumber
import json
import os
import re
import sys

# ── Config ────────────────────────────────────────────────────────────────────

def load_config(path="config.json"):
    with open(path) as f:
        cfg = json.load(f)
    _set_derived_paths(cfg, cfg["tax_year"])
    return cfg

def _set_derived_paths(cfg, year):
    cfg["pdf_dir"]  = os.path.join(cfg["base_dir"], year)
    cfg["dist_dir"] = os.path.join(cfg["base_dir"], year, "distributions")
    cfg["assets_dir"] = os.path.join(cfg["base_dir"], year, "assets")
    cfg["output_txt"] = os.path.join(cfg["base_dir"], year, f"T3_results_{year}.txt")

def parse_args():
    parser = argparse.ArgumentParser(description="Parse CDS T3 PDFs and output distribution JSON + ACB Excel.")
    parser.add_argument("--config", default="config.json",  help="Path to config.json")
    parser.add_argument("--year",   help="Override tax year in config")
    parser.add_argument("--funds",  nargs="+", help="Override fund list (e.g. --funds VBAL ZCN)")
    return parser.parse_args()

def load_funds_meta(config_path):
    """Load funds.json from same directory as config.json. Returns {} if not found."""
    funds_meta_path = os.path.join(os.path.dirname(os.path.abspath(config_path)), "funds.json")
    if os.path.exists(funds_meta_path):
        with open(funds_meta_path) as f:
            return json.load(f)
    return {}

# ── T3 box → JSON field mapping ───────────────────────────────────────────────

BOX_TO_FIELD = {
    "21": "capitalGains",
    "49": "actualAmountOfEligibleDividends",
    "23": "actualAmountOfNonEligibleDividends",
    "25": "foreignNonBusinessIncome",
    "42": "returnOfCapital",
    "34": "foreignNonBusinessIncomeTaxPaid",
    # "G" Other Income is handled separately (label in the PDF is the single letter "G")
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def is_number(s):
    try:    float(s); return True
    except: return False

def is_date(s):
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}$", s or ""))

def fmt(v):
    """Full-precision decimal display — no scientific notation."""
    s = f"{v:.9g}"
    if "e" in s or "E" in s:
        s = f"{v:.9f}".rstrip("0").rstrip(".")
    return s

# ── Core extractor ────────────────────────────────────────────────────────────

def extract_distributions(pdf_path, fund_name, calc_method_override=None):
    """
    Parse a CDS T3 PDF and return a dict:
        {"name": fund_name, "distributions": [...]}

    Each distribution dict contains recordDate, paymentDate, total,
    component fields, optional capitalGainsDistributedAsCash flag,
    and private _cashAmount / _nonCashAmount fields (stripped before JSON output).

    calc_method_override: "RATE", "PERCENT", or None (auto-detect from PDF).
    """
    with pdfplumber.open(pdf_path) as pdf:
        words = pdf.pages[0].extract_words()

    # Group words into rows with 2px tolerance to handle sub-pixel PDF variation
    rows = {}
    for w in words:
        top = w["top"]
        key = next((k for k in rows if abs(k - top) <= 1), round(top, 0))
        rows.setdefault(key, []).append(w)
    for top in rows:
        rows[top].sort(key=lambda w: w["x0"])
    sorted_tops = sorted(rows.keys())

    # ── Detect RATE vs PERCENT mode ───────────────────────────────────────────
    if calc_method_override:
        calc_method = calc_method_override.upper()
    else:
        calc_method = "RATE"
        for top in sorted_tops:
            txt = " ".join(w["text"] for w in rows[top])
            if "CALCULATION METHOD" in txt and "PERCENT" in txt:
                calc_method = "PERCENT"
                break

    # ── Find distribution column x-positions ─────────────────────────────────
    dist_x = []
    for top in sorted_tops:
        rw    = rows[top]
        texts = [w["text"] for w in rw]
        if "Distribution" in texts and any(t.isdigit() for t in texts):
            i = 0
            while i < len(rw):
                if rw[i]["text"] == "Distribution" and i+1 < len(rw) and rw[i+1]["text"].isdigit():
                    dist_x.append(rw[i]["x0"]); i += 2
                else:
                    i += 1
            if dist_x:
                break

    if not dist_x:
        return {"name": fund_name, "distributions": []}

    COL_TOL       = 35
    LEFT_BOUNDARY = dist_x[0] - 10

    def col_of(x0):
        for i, cx in enumerate(dist_x):
            if abs(x0 - cx) <= COL_TOL:
                return i
        return None

    def values_on_row(top):
        out = {}
        for w in rows.get(top, []):
            if w["x0"] >= LEFT_BOUNDARY:
                c = col_of(w["x0"])
                if c is not None and (is_date(w["text"]) or is_number(w["text"])):
                    out[c] = w["text"]
        return out

    def find_values_near(label_top, window=5):
        for t2 in sorted_tops:
            if abs(t2 - label_top) <= window:
                v = values_on_row(t2)
                if v:
                    return v
        return {}

    raw = {}  # field → {col_index: value_str}

    # Track cash and non-cash rows for phantom detection
    cash_vals     = {}   # col → float
    non_cash_vals = {}   # col → float

    for top in sorted_tops:
        lw = [w for w in rows[top] if w["x0"] < LEFT_BOUNDARY]
        if not lw:
            continue
        label = " ".join(w["text"] for w in lw).strip()
        parts = label.split()

        # Box-number rows
        for box, field in BOX_TO_FIELD.items():
            if parts[0] == box:
                v = values_on_row(top)
                if v:
                    raw[field] = v
                break

        # "G" row (Other Income)
        if label == "G":
            v = values_on_row(top)
            if v:
                raw["otherIncome"] = v

        # Total Income row
        if "Total Income ($) per unit being" in label:
            v = find_values_near(top, window=5)
            if v:
                raw["total"] = v

        # Total Cash Distribution row
        if "Total Cash Distribution" in label and "Non" not in label:
            for t2 in sorted_tops:
                if abs(t2 - top) <= 5:
                    for w in rows.get(t2, []):
                        if w["x0"] >= LEFT_BOUNDARY and is_number(w["text"]):
                            c = col_of(w["x0"])
                            if c is not None:
                                cash_vals[c] = float(w["text"])

        # Total Non-Cash Distribution row
        if "Total Non Cash Distribution" in label:
            for t2 in sorted_tops:
                if abs(t2 - top) <= 5:
                    for w in rows.get(t2, []):
                        if w["x0"] >= LEFT_BOUNDARY and is_number(w["text"]):
                            c = col_of(w["x0"])
                            if c is not None:
                                non_cash_vals[c] = float(w["text"])

        # Record dates
        if re.search(r"R\s+e\s+c\s+o\s+r\s+d", label):
            v = find_values_near(top, window=5)
            if v:
                raw["recordDate"] = v

        # Payment dates
        if re.search(r"P\s+a\s+y\s+m\s+e\s+n\s+t", label):
            v = find_values_near(top, window=5)
            if v:
                raw["paymentDate"] = v

    # ── Assemble one distribution dict per column ─────────────────────────────
    distributions = []
    for col in range(len(dist_x)):
        dist = {}
        for field, col_vals in raw.items():
            if col in col_vals:
                v = col_vals[col]
                dist[field] = v if is_date(v) else float(v)

        if "recordDate" not in dist or "total" not in dist:
            continue

        # PERCENT mode: convert percentages → dollar amounts
        if calc_method == "PERCENT":
            total = dist["total"]
            for field in list(dist.keys()):
                if field not in ("recordDate", "paymentDate", "total"):
                    dist[field] = dist[field] / 100.0 * total

        # Drop zero fields
        if dist.get("capitalGains", 0) == 0:
            dist.pop("capitalGains", None)
        if dist.get("actualAmountOfNonEligibleDividends", 0) == 0:
            dist.pop("actualAmountOfNonEligibleDividends", None)

        # capitalGainsDistributedAsCash flag
        cash_total = cash_vals.get(col, 0.0)
        if dist.get("capitalGains") and cash_total and abs(cash_total - dist["total"]) < 1e-9:
            dist["capitalGainsDistributedAsCash"] = True

        # Phantom (non-cash) distribution detection
        non_cash = non_cash_vals.get(col, 0.0)
        if non_cash > 0:
            total_val = dist["total"]
            cash = cash_vals.get(col, 0.0)
            if calc_method == "PERCENT":
                non_cash = (non_cash / 100.0 * total_val) if non_cash > 1 else non_cash
                cash     = (cash     / 100.0 * total_val) if cash > 1     else cash
            dist["_cashAmount"]    = cash
            dist["_nonCashAmount"] = non_cash

        distributions.append(dist)

    return {"name": fund_name, "distributions": distributions}


# ── Excel ACB helper output ───────────────────────────────────────────────────

def write_acb_excel(result, output_path):
    """
    Writes a 3-column Excel file: Date | Transaction | Price/Share
      - Regular ROC row : Price/Share = +returnOfCapital per unit  (reduces ACB)
      - Phantom ROC row : Price/Share = -nonCashAmount per unit    (increases ACB)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = result["name"]

    header_font  = Font(name="Arial", bold=True, size=10)
    data_font    = Font(name="Arial", size=10)
    thin_side    = Side(style="thin", color="CCCCCC")
    thin_border  = Border(bottom=thin_side)
    center_align = Alignment(horizontal="center")
    right_align  = Alignment(horizontal="right")

    headers = ["Date", "Transaction", "Price / Share"]
    widths  = [14, 14, 16]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = w

    grey = PatternFill("solid", start_color="D9D9D9")
    for col in range(1, 4):
        ws.cell(row=1, column=col).fill = grey

    row_num = 2
    for d in result["distributions"]:
        date_raw = d.get("recordDate", "")
        roc      = d.get("returnOfCapital", 0)
        non_cash = d.get("_nonCashAmount", 0)

        try:
            dt = datetime.strptime(date_raw, "%Y-%m-%d")
            date_str = dt.strftime("%Y-%b-%d")
        except Exception:
            date_str = date_raw

        def write_row(date_str, price):
            nonlocal row_num
            cells = [
                ws.cell(row=row_num, column=1, value=date_str),
                ws.cell(row=row_num, column=2, value="ROC"),
                ws.cell(row=row_num, column=3, value=price),
            ]
            cells[0].alignment = center_align
            cells[1].alignment = center_align
            cells[2].alignment = right_align
            cells[2].number_format = '"$"#,##0.#########;"-$"#,##0.#########'
            for cell in cells:
                cell.font = data_font
                cell.border = thin_border
            row_num += 1

        if roc:
            write_row(date_str, roc)
        if non_cash:
            write_row(date_str, -non_cash)

    wb.save(output_path)


# ── Custom JSON serializer ────────────────────────────────────────────────────

def to_json(obj, indent=0):
    """Serialize to JSON with full float precision and no scientific notation."""
    pad = "  " * indent
    if isinstance(obj, dict):
        lines = ["{"]
        items = list(obj.items())
        for i, (k, v) in enumerate(items):
            comma = "," if i < len(items) - 1 else ""
            if isinstance(v, bool):
                lines.append(f"{pad}  {json.dumps(k)}: {str(v).lower()}{comma}")
            elif isinstance(v, float):
                s = repr(v)
                if "e" in s or "E" in s:
                    s = f"{v:.9f}".rstrip("0").rstrip(".")
                lines.append(f"{pad}  {json.dumps(k)}: {s}{comma}")
            elif isinstance(v, (dict, list)):
                lines.append(f"{pad}  {json.dumps(k)}: {to_json(v, indent+1)}{comma}")
            else:
                lines.append(f"{pad}  {json.dumps(k)}: {json.dumps(v)}{comma}")
        lines.append(f"{pad}}}")
        return "\n".join(lines)
    elif isinstance(obj, list):
        if not obj:
            return "[]"
        parts = []
        for i, v in enumerate(obj):
            comma = "," if i < len(obj) - 1 else ""
            parts.append(f"{pad}  {to_json(v, indent+1)}{comma}")
        return "[\n" + "\n".join(parts) + f"\n{pad}]"
    else:
        return json.dumps(obj)


# ── XLS extraction (prior years) ─────────────────────────────────────────────

# Fixed row positions in the CDS T3 Excel template (1-indexed, openpyxl convention)
_XLS_ROW_TOTAL     = 19
_XLS_ROW_RECORD    = 20
_XLS_ROW_PAYMENT   = 21
_XLS_ROW_CASH      = 22
_XLS_ROW_NONCASH   = 23
_XLS_ROW_FIELDS = {
    # row: (json_field, description)
    25: ("capitalGains",                          "Capital gain"),
    26: ("actualAmountOfEligibleDividends",        "Eligible dividends"),
    27: ("actualAmountOfNonEligibleDividends",     "Non-eligible dividends"),
    28: ("foreignBusinessIncome",                  "Foreign business income"),
    29: ("foreignNonBusinessIncome",               "Foreign non-business income"),
    30: ("otherIncome",                            "Other income"),
    32: ("returnOfCapital",                        "Return of capital"),
    34: ("capitalGainsEligibleForDeduction",       "Cap gains eligible for deduction"),
    35: ("foreignBusinessIncomeTaxPaid",           "Foreign business income tax paid"),
    36: ("foreignNonBusinessIncomeTaxPaid",        "Foreign non-business income tax paid"),
}
_XLS_COL_FIRST_DIST = 4   # column D = index 4 (1-based), up to 14 distributions


def extract_distributions_from_xls(xls_path, fund_name, calc_method_override=None):
    """
    Parse a CDS T3 Excel (.xls) file (tax years 2024 and earlier) and return
    the same dict structure as extract_distributions():
        {
            "name": fund_name,
            "distributions": [ { recordDate, paymentDate, total, <box fields> }, ... ]
        }

    Requires: pip install xlrd
    xlrd reads the binary OLE2 .xls format directly — no conversion needed.

    Row layout (1-based, matching the CDS template):
        Row 15 col G : calculation method (1=PERCENT, 2=RATE)
        Row 19       : total distribution per unit
        Row 20       : record dates
        Row 21       : payment dates
        Row 22       : total cash distribution per unit
        Row 23       : total non-cash distribution per unit
        Rows 25–36   : T3 box values (see _XLS_ROW_FIELDS)
        Cols D–Q     : one column per distribution (up to 14)

    Calculation methods:
        RATE    (method=2): values are per-unit dollar amounts — use directly.
        PERCENT (method=1): values are percentages of total — multiply by total/100.

    calc_method_override: "RATE" or "PERCENT" — overrides the spreadsheet value.
    """
    try:
        import xlrd
    except ImportError:
        raise ImportError(
            "xlrd is required to parse CDS T3 XLS files (tax years 2024 and earlier).\n"
            "Install it with:  pip install xlrd"
        )

    SHEET_NAME   = "T3, R16"
    # xlrd uses 0-based row/col indices
    # openpyxl-style (row, col) 1-based → xlrd (row-1, col-1)
    ROW_TOTAL    = _XLS_ROW_TOTAL   - 1   # 18
    ROW_RECORD   = _XLS_ROW_RECORD  - 1   # 19
    ROW_PAYMENT  = _XLS_ROW_PAYMENT - 1   # 20
    ROW_CASH     = _XLS_ROW_CASH    - 1   # 21
    ROW_NONCASH  = _XLS_ROW_NONCASH - 1   # 22
    COL_METHOD   = 6                       # col G = index 6 (0-based)
    COL_FIRST    = _XLS_COL_FIRST_DIST - 1  # col D = index 3

    wb = xlrd.open_workbook(xls_path)

    sheet_names = wb.sheet_names()
    if SHEET_NAME not in sheet_names:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found in {xls_path}. "
            f"Available sheets: {sheet_names}"
        )
    ws = wb.sheet_by_name(SHEET_NAME)

    def cell(row0, col0):
        """Return cell value using 0-based indices. Returns None for empty cells."""
        try:
            c = ws.cell(row0, col0)
            if c.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                return None
            if c.ctype == xlrd.XL_CELL_TEXT and c.value.strip() == "":
                return None
            return c.value
        except IndexError:
            return None

    def cell_float(row0, col0):
        """Return cell value as float, or None if empty/non-numeric."""
        v = cell(row0, col0)
        if v is None:
            return None
        try:
            f = float(v)
            return f if f != 0.0 else None
        except (TypeError, ValueError):
            return None

    def parse_xls_date(v):
        """Parse a date cell — xlrd may return a float (date serial) or string."""
        if v is None:
            return None
        # xlrd sometimes returns date serials as floats
        if isinstance(v, float):
            try:
                t = xlrd.xldate_as_tuple(v, wb.datemode)
                return f"{t[0]:04d}-{t[1]:02d}-{t[2]:02d}"
            except Exception:
                pass
        s = str(v).strip()
        for fmt_str in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                from datetime import datetime as _dt
                return _dt.strptime(s, fmt_str).strftime("%Y-%m-%d")
            except ValueError:
                pass
        return s   # return as-is if unparseable

    # ── Determine calculation method ──────────────────────────────────────────
    # Row 15, col G (0-based: row 14, col 6)
    raw_method = cell(14, COL_METHOD)
    if calc_method_override:
        method = calc_method_override.upper()
    elif raw_method == 1 or raw_method == 1.0:
        method = "PERCENT"
    elif raw_method == 2 or raw_method == 2.0:
        method = "RATE"
    else:
        s = str(raw_method).upper() if raw_method else ""
        if "PERCENT" in s or "PER CENT" in s or "%" in s:
            method = "PERCENT"
        else:
            method = "RATE"   # safe default

    distributions = []

    # ── Iterate distribution columns D through Q (up to 14) ──────────────────
    for col0 in range(COL_FIRST, COL_FIRST + 14):
        record_date_raw  = cell(ROW_RECORD,  col0)
        payment_date_raw = cell(ROW_PAYMENT, col0)
        total_raw        = cell_float(ROW_TOTAL, col0)

        if record_date_raw is None or total_raw is None:
            break   # no more distributions

        record_date  = parse_xls_date(record_date_raw)
        payment_date = parse_xls_date(payment_date_raw)
        total        = total_raw

        non_cash = cell_float(ROW_NONCASH, col0) or 0.0
        cash_val = cell_float(ROW_CASH,    col0)
        cash     = cash_val if cash_val is not None else total

        dist = {
            "recordDate":  record_date,
            "paymentDate": payment_date,
            "total":       round(total, 7),
        }

        # ── Parse T3 box fields ───────────────────────────────────────────────
        for row1, (field, _) in _XLS_ROW_FIELDS.items():
            value = cell_float(row1 - 1, col0)   # convert 1-based to 0-based
            if value is None:
                continue
            if method == "PERCENT":
                # Value is a percentage (e.g. 93.39834 = 93.39834% of total)
                value = round(total * value / 100.0, 7)
            else:
                value = round(value, 7)
            dist[field] = value

        # ── Tag phantom (non-cash) capital gains ──────────────────────────────
        cap_gains = dist.get("capitalGains", 0.0)
        if non_cash > 0.0 and cap_gains > 0.0:
            if abs(cash) < 0.0001:
                # Fully phantom — no cash received for the cap gains portion
                dist["nonCashCapitalGains"] = round(cap_gains, 7)
            elif non_cash < total:
                # Partially phantom
                dist["nonCashCapitalGains"] = round(non_cash, 7)
                if non_cash < cap_gains:
                    dist["capitalGainsDistributedAsCash"] = False
        elif cap_gains > 0.0 and non_cash == 0.0:
            # Fully cash — cap gains distributed as cash
            dist["capitalGainsDistributedAsCash"] = True

        distributions.append(dist)

    return {"name": fund_name, "distributions": distributions}


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    args = parse_args()
    cfg  = load_config(args.config)

    # Apply CLI overrides
    if args.year:
        cfg["tax_year"] = args.year
        _set_derived_paths(cfg, args.year)

    funds    = args.funds if args.funds else cfg["funds"]
    tax_year = cfg["tax_year"]
    pdf_dir  = cfg["pdf_dir"]
    dist_dir = cfg["dist_dir"]

    os.makedirs(dist_dir, exist_ok=True)

    # Load per-fund metadata (calc_method_override etc.)
    funds_meta = load_funds_meta(args.config)

    # CDS switched from XLS to PDF format starting with tax year 2025
    use_pdf = int(tax_year) >= 2025

    print(f"Tax year : {tax_year}")
    print(f"Format   : {'PDF (2025+)' if use_pdf else 'XLS (2024 and earlier)'}")
    print(f"Src dir  : {pdf_dir}")
    print(f"Dist dir : {dist_dir}")
    print()

    skipped = []

    for fund in funds:
        override = funds_meta.get(fund, {}).get("calc_method_override")

        if use_pdf:
            src_path = os.path.join(pdf_dir, f"{fund}_T3_{tax_year}.pdf")
            src_type = "PDF"
        else:
            src_path = os.path.join(pdf_dir, f"{fund}_T3_{tax_year}.xls")
            src_type = "XLS"

        if not os.path.exists(src_path):
            print(f"  WARNING: {src_type} not found — skipping {fund}")
            print(f"           Expected : {src_path}")
            print(f"           Download from: https://ctbsext.posttrade.cds.ca/ctbsExt/")
            skipped.append(fund)
            continue

        print(f"Processing {fund} ({src_type})..." + (f" [override: {override}]" if override else ""))

        try:
            if use_pdf:
                result = extract_distributions(src_path, fund, calc_method_override=override)
            else:
                result = extract_distributions_from_xls(src_path, fund, calc_method_override=override)
        except Exception as e:
            print(f"  WARNING: Failed to parse {src_type} for {fund}: {e}")
            skipped.append(fund)
            continue

        if not result["distributions"]:
            print(f"  WARNING: No distributions found for {fund} — file structure may differ")
            print(f"           Check the file is the CDS T3 statement, not the broker version.")
            print(f"           Try setting calc_method_override in funds.json to 'RATE' or 'PERCENT'.")
            skipped.append(fund)
            continue
        else:
            print(f"  Found {len(result['distributions'])} distribution(s)")
            for d in result["distributions"]:
                date     = d.get("recordDate", "")
                total    = d.get("total", 0)
                non_cash = d.get("_nonCashAmount", 0)

                parts = []
                for label, key in [
                    ("capGains", "capitalGains"),
                    ("eligDiv",  "actualAmountOfEligibleDividends"),
                    ("forInc",   "foreignNonBusinessIncome"),
                    ("other",    "otherIncome"),
                    ("ROC",      "returnOfCapital"),
                    ("forTax",   "foreignNonBusinessIncomeTaxPaid"),
                ]:
                    v = d.get(key, 0)
                    if v:
                        parts.append(f"{label}={fmt(v)}")

                breakdown = "  [" + "  ".join(parts) + "]" if parts else ""
                phantom   = f"   nonCash={fmt(non_cash)}  ACB+={fmt(non_cash)}" if non_cash else ""
                print(f"     {date}  total={total}{breakdown}{phantom}")
            print()

        # Strip internal tracking fields before writing JSON
        # _nonCashAmount is saved as nonCashCapitalGains for use by update_acb.py
        def clean_dist(d):
            out = {k: v for k, v in d.items() if not k.startswith("_")}
            if d.get("_nonCashAmount", 0) and "nonCashCapitalGains" not in out:
                out["nonCashCapitalGains"] = d["_nonCashAmount"]
            return out

        clean = {
            "name": result["name"],
            "distributions": [clean_dist(d) for d in result["distributions"]]
        }

        json_path = os.path.join(dist_dir, f"{fund}.json")
        with open(json_path, "w") as f:
            f.write(to_json(clean))
        print(f"  JSON  → {json_path}")

        xlsx_path = os.path.join(dist_dir, f"{fund}_ACB_{tax_year}.xlsx")
        write_acb_excel(result, xlsx_path)
        print(f"  Excel → {xlsx_path}\n")



    if skipped:
        print(f"\n  WARNING: {len(skipped)} fund(s) skipped: {', '.join(skipped)}")
        print(f"  Re-run Step 1 after downloading or fixing the missing files.")
        sys.exit(2)   # non-zero so run_t3.py can detect warnings


if __name__ == "__main__":
    main()

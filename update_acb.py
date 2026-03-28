"""
update_acb.py
-------------
Step 2 of the T3 pipeline.

Reads distribution JSONs produced by parse_t3_pdfs.py (Step 1) and inserts
ROC rows directly into the ACB spreadsheet in date order, one fund at a time.

Safety features:
  - Timestamped backup created before any changes
  - Duplicate detection: date + ROC + amount check
  - Year boundary: only rows within tax_year are inserted
  - Formula integrity check after every insertion
  - Automatic rollback on any failure
  - Dry-run mode (--dry-run) to preview without touching the file

CLI (via run_t3.py --step 2, or directly):
  python update_acb.py
  python update_acb.py --config config.json --year 2025
  python update_acb.py --funds VBAL ZCN
  python update_acb.py --dry-run
"""

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime, date
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from t3_colors import c, BOLD, DIM, CYAN, GREEN, YELLOW, RED
except ImportError:
    def c(code, text): return text
    BOLD = DIM = CYAN = GREEN = YELLOW = RED = ""

# ── Formatting constants (must match acb_worksheet_template) ─────────────────
DATE_FMT    = "YYYY-MMM-DD"
PRICE_FMT   = '$#,##0.#######;-$#,##0.#######'
SHARES_FMT  = '#,##0.####'
COMM_FMT    = '$#,##0.00;-$#,##0.00;"-"'
CAPGAIN_FMT = '$#,##0.00;-$#,##0.00;$0.00'
BALANCE_FMT = '#,##0.####'
ACBCHG_FMT  = '$#,##0.00;-$#,##0.00;$0.00'
ACB_FMT     = '$#,##0.00;-$#,##0.00;$0.00'
ACBSHR_FMT  = '$#,##0.######;"-$"#,##0.######'

INPUT_COLOR  = "000000"
CALC_COLOR   = "000000"
HEADER_COLOR = "1F4E79"
ALT_BG       = None          # no alternating row background on inserted rows

thin_black = Side(style="thin", color="000000")
BORDER = Border(bottom=thin_black, top=thin_black, left=thin_black, right=thin_black)

# ── Config loading ────────────────────────────────────────────────────────────

def load_config(config_path, year_override=None, funds_override=None):
    with open(config_path) as f:
        cfg = json.load(f)
    year = year_override or cfg["tax_year"]
    cfg["year"] = year
    cfg["dist_dir"] = os.path.join(cfg["base_dir"], year, "distributions")
    if funds_override:
        cfg["funds"] = funds_override
    return cfg

# ── Date parsing ──────────────────────────────────────────────────────────────

MONTH_MAP = {
    "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
    "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12
}

def parse_date(val):
    """Parse a date from various formats into a datetime.date object."""
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    # YYYY-MMM-DD
    m = re.match(r'(\d{4})-([A-Za-z]{3})-(\d{2})', s)
    if m:
        return date(int(m.group(1)), MONTH_MAP[m.group(2).lower()], int(m.group(3)))
    # YYYY-MM-DD
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    raise ValueError(f"Cannot parse date: {val!r}")

def format_date_str(d):
    """Format a date as YYYY-MMM-DD string for Excel."""
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{d.year}-{months[d.month-1]}-{d.day:02d}"

# ── Stylesheet helpers ────────────────────────────────────────────────────────

def _font(color, bold=False):
    return Font(name="Arial", size=10, color=color, bold=bold)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)

def _alt_fill(row_number):
    return None  # inserted rows inherit no background — matches existing sheet style

def _apply(cell, fmt, font, align, fill):
    cell.number_format = fmt
    cell.font          = font
    cell.alignment     = align
    cell.border        = BORDER
    if fill:
        cell.fill = fill

# ── Formula writer (identical logic to template builder) ─────────────────────

def write_formulas(ws, r, fill):
    """Write formula cells F-J for row r using chained references."""
    right  = Alignment(horizontal="right",  vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    prev   = r - 1
    G_prev = f"G{prev}" if r > 2 else "0"
    I_prev = f"I{prev}" if r > 2 else "0"
    J_prev = f"J{prev}" if r > 2 else "0"

    # G — Share Balance
    g = (f'=ROUND(IF(B{r}="Buy",{G_prev}+D{r},'
         f'IF(B{r}="Sell",{G_prev}-D{r},{G_prev})),4)')
    _apply(ws.cell(r, 7, g), BALANCE_FMT, _font(CALC_COLOR), right, fill)

    # H — ACB Change
    h = (f'=IF(B{r}="Buy",'
         f'ROUND(C{r}*D{r}+IF(ISNUMBER(E{r}),E{r},0),2),'
         f'IF(B{r}="Sell",'
         f'ROUND(-({J_prev}*D{r}),2),'
         f'IF(B{r}="ROC",'
         f'ROUND(-C{r}*{G_prev},2),0)))')
    _apply(ws.cell(r, 8, h), ACBCHG_FMT, _font(CALC_COLOR), right, fill)

    # I — Cumulative ACB
    i = f'=ROUND({I_prev}+H{r},2)'
    _apply(ws.cell(r, 9, i), ACB_FMT, _font(CALC_COLOR), right, fill)

    # F — Capital Gains
    f_ = (f'=IF(B{r}="Sell",'
          f'ROUND((C{r}-{J_prev})*D{r}-IF(ISNUMBER(E{r}),E{r},0),2),0)')
    _apply(ws.cell(r, 6, f_), CAPGAIN_FMT, _font(CALC_COLOR), right, fill)

    # J — ACB / Share
    j = f'=IF(G{r}<>0,ROUND(I{r}/G{r},7),0)'
    _apply(ws.cell(r, 10, j), ACBSHR_FMT, _font(CALC_COLOR), right, fill)

def rewrite_formulas_from(ws, start_row, last_data_row):
    """Rewrite all formula rows from start_row downward after an insertion."""
    for r in range(start_row, last_data_row + 1):
        write_formulas(ws, r, fill=None)

# ── Formula integrity check ───────────────────────────────────────────────────

FORMULA_PATTERNS = {
    6:  r'=IF\(B\d+="Sell"',          # Capital Gains
    7:  r'=ROUND\(IF\(B\d+="Buy"',    # Share Balance
    8:  r'=IF\(B\d+="Buy"',           # ACB Change
    9:  r'=ROUND\(',                   # ACB
    10: r'=IF\(G\d+<>0',              # ACB / Share
}

def check_formula_integrity(ws, start_row, last_data_row):
    """Verify formula cells F-J exist and match expected patterns."""
    errors = []
    for r in range(start_row, last_data_row + 1):
        txn = ws.cell(r, 2).value
        if not txn:
            continue
        for col, pattern in FORMULA_PATTERNS.items():
            val = ws.cell(r, col).value
            if val is None:
                errors.append(f"Row {r} col {get_column_letter(col)}: formula missing")
            elif not str(val).startswith("="):
                errors.append(f"Row {r} col {get_column_letter(col)}: not a formula: {val!r}")
            elif not re.match(pattern, str(val)):
                errors.append(f"Row {r} col {get_column_letter(col)}: unexpected formula: {val!r}")
    return errors

# ── Sheet helpers ─────────────────────────────────────────────────────────────

def find_last_data_row(ws):
    """Find the last row with a transaction value in col B."""
    last = 1
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if val and str(val).strip() in ("Buy", "Sell", "ROC"):
            last = r
    return last

def get_sheet_roc_rows(ws, last_data_row):
    """Return set of (date_obj, amount) tuples for existing ROC rows."""
    existing = set()
    for r in range(2, last_data_row + 1):
        txn = ws.cell(r, 2).value
        if txn and str(txn).strip() == "ROC":
            try:
                d   = parse_date(ws.cell(r, 1).value)
                amt = ws.cell(r, 3).value
                if amt is not None:
                    existing.add((d, round(float(amt), 7)))
            except Exception:
                pass
    return existing

def find_insertion_row(ws, target_date, last_data_row):
    """
    Find the row number where a new row with target_date should be inserted.
    Inserts AFTER the last existing row with the same date (to keep same-date
    entries together), or after the last row with an earlier date.
    Returns last_data_row + 1 if it belongs at the end.
    """
    insert_after = 1  # default: after header
    for r in range(2, last_data_row + 1):
        try:
            d = parse_date(ws.cell(r, 1).value)
            if d <= target_date:
                insert_after = r
        except Exception:
            pass
    return insert_after + 1  # insert BEFORE this row number

# ── Backup ────────────────────────────────────────────────────────────────────

def make_backup(acb_path):
    ts      = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    base    = os.path.splitext(acb_path)[0]
    backup  = f"{base}_{ts}.bak.xlsx"
    shutil.copy2(acb_path, backup)
    return backup

# ── ROC row builder ───────────────────────────────────────────────────────────

def build_roc_rows(dist_json, year):
    """
    Extract ROC rows from a distribution JSON for a given year.
    Returns list of (date_obj, amount) tuples sorted by date.

    Two row types are produced per distribution (when applicable):
      - Regular ROC   : positive amount  — reduces ACB
      - Phantom ROC   : negative amount  — capitalGains not distributed as cash,
                        increases ACB even though no cash was received
    """
    roc_rows = []
    year_int = int(year)

    for entry in dist_json.get("distributions", []):
        record_date_str = entry.get("recordDate") or entry.get("record_date", "")
        try:
            d = parse_date(record_date_str)
        except Exception:
            continue
        if d.year != year_int:
            continue

        # Regular ROC row (positive price — reduces ACB)
        roc_amount = entry.get("returnOfCapital", 0) or 0
        if roc_amount != 0:
            roc_rows.append((d, round(float(roc_amount), 7)))

        # Phantom ROC row (negative price — increases ACB, no cash received)
        # Use nonCashCapitalGains if present (partial phantom, e.g. VRE Dec-30)
        # otherwise fall back to full capitalGains if not distributed as cash
        non_cash_cg = entry.get("nonCashCapitalGains", 0) or 0
        cap_gains   = entry.get("capitalGains", 0) or 0
        if non_cash_cg != 0:
            roc_rows.append((d, round(-float(non_cash_cg), 7)))
        elif cap_gains != 0 and not entry.get("capitalGainsDistributedAsCash", False):
            roc_rows.append((d, round(-float(cap_gains), 7)))

    # Sort by date, regular ROC before phantom on same date (positive before negative)
    roc_rows.sort(key=lambda x: (x[0], x[1] < 0))
    return roc_rows

# ── Core insertion logic ──────────────────────────────────────────────────────

def insert_roc_rows(ws, roc_rows, existing_roc, year, dry_run, fund):
    """
    Insert ROC rows into the worksheet.
    Returns (inserted, skipped_dup, warned, errors) counts.
    """
    right  = Alignment(horizontal="right",  vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    inserted    = 0
    skipped_dup = 0
    warned      = 0
    errors      = []
    year_int    = int(year)

    for target_date, amount in roc_rows:
        # Year boundary check
        if target_date.year != year_int:
            continue

        date_str = format_date_str(target_date)

        # Duplicate check
        key = (target_date, round(amount, 7))

        # Exact duplicate — already in sheet with same date AND same amount
        if key in existing_roc:
            print(c(DIM,    f"  [{fund}] SKIP (duplicate): {date_str}  ROC {amount:+.7f}"))
            skipped_dup += 1
            continue

        # Same date, same sign, different amount — genuinely ambiguous, warn and skip
        # (positive+negative on the same date is normal: regular ROC + phantom)
        is_negative = amount < 0
        same_sign_same_date = [
            a for (d, a) in existing_roc
            if d == target_date and (a < 0) == is_negative
        ]
        if same_sign_same_date:
            print(c(YELLOW, f"  [{fund}] WARNING: ROC row for {date_str} already exists "
                  f"with same sign but different amount(s) {same_sign_same_date} vs {amount}. "
                  f"Skipping — please review manually."))
            warned += 1
            continue

        if dry_run:
            print(c(CYAN,   f"  [{fund}] DRY-RUN — would insert: {date_str}  ROC {amount:+.7f}"))
            inserted += 1
            continue

        # Find insertion point
        last_data_row  = find_last_data_row(ws)
        insert_at      = find_insertion_row(ws, target_date, last_data_row)

        # Insert blank row
        ws.insert_rows(insert_at)

        # New last data row after insertion
        new_last = last_data_row + 1

        # Write input values (cols A-E) — black text, no background, matches existing rows
        font   = _font(INPUT_COLOR)

        date_cell = ws.cell(insert_at, 1, date_str)
        _apply(date_cell, DATE_FMT, font, left, None)

        txn_cell = ws.cell(insert_at, 2, "ROC")
        _apply(txn_cell, "@", font, left, None)

        amt_cell = ws.cell(insert_at, 3, amount)
        _apply(amt_cell, PRICE_FMT, font, right, None)

        # Cols D and E — blank but formatted (Shares uses General so blank ROC rows show nothing)
        for col, fmt in [(4, 'General'), (5, COMM_FMT)]:
            _apply(ws.cell(insert_at, col), fmt, font, right, None)

        # Write formulas for inserted row and all rows below
        rewrite_formulas_from(ws, insert_at, new_last)

        # Update existing_roc set so subsequent iterations see this new row
        existing_roc.add(key)

        print(c(GREEN,  f"  [{fund}] INSERTED: {date_str}  ROC {amount:+.7f}  at row {insert_at}"))
        inserted += 1

    return inserted, skipped_dup, warned, errors

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Step 2 — Insert ROC rows from T3 pipeline into ACB spreadsheet.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python update_acb.py                        Interactive (uses config.json)
  python update_acb.py --year 2025            Override tax year
  python update_acb.py --funds VBAL ZCN       Process specific funds only
  python update_acb.py --dry-run              Preview changes without modifying file
        """
    )
    parser.add_argument("--config",  default="config.json")
    parser.add_argument("--year",    default=None)
    parser.add_argument("--funds",   nargs="+", default=None)
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview what would be inserted without modifying the file")
    args = parser.parse_args()

    # ── Load config ───────────────────────────────────────────────────────────
    cfg = load_config(args.config, args.year, args.funds)
    year     = cfg["year"]
    funds    = cfg["funds"]
    acb_path = cfg["acb_spreadsheet"]
    dist_dir = cfg["dist_dir"]

    print()
    print(c(DIM,    f"  Tax year   : {year}"))
    print(c(DIM,    f"  Funds      : {', '.join(funds)}"))
    print(c(DIM,    f"  ACB file   : {acb_path}"))
    print(c(DIM,    f"  Dist dir   : {dist_dir}"))
    if args.dry_run:
        print(c(YELLOW, "  Mode       : DRY-RUN (no changes will be made)"))
    print()

    # ── Validate paths ────────────────────────────────────────────────────────
    if not os.path.exists(acb_path):
        print(c(RED, f"  ERROR: ACB spreadsheet not found: {acb_path}"))
        sys.exit(1)
    if not os.path.isdir(dist_dir):
        print(c(RED, f"  ERROR: Distributions folder not found: {dist_dir}"))
        print(c(DIM,  "         Run Step 1 first to generate distribution JSONs."))
        sys.exit(1)

    # ── Year boundary warning ─────────────────────────────────────────────────
    current_year = datetime.now().year
    if int(year) < current_year - 1:
        print(c(YELLOW, f"WARNING: You are inserting ROC rows for {year} into a spreadsheet"))
        print(c(YELLOW, f"         that likely contains transactions from later years."))
        print(c(YELLOW, f"         Duplicate checks will run. Rows outside {year} will not be touched."))
        confirm = input(f"  Proceed? [Y/N]: ").strip().upper()
        if confirm != "Y":
            print(c(YELLOW, "Aborted."))
            sys.exit(0)
        print()

    # ── Backup ────────────────────────────────────────────────────────────────
    backup_path = None
    if not args.dry_run:
        backup_path = make_backup(acb_path)
        print(c(DIM, f"  Backup created: {os.path.basename(backup_path)}"))
        print()

    # ── Load workbook ─────────────────────────────────────────────────────────
    wb = load_workbook(acb_path)

    total_inserted = 0
    total_skipped  = 0
    total_warned   = 0
    rollback       = False
    rollback_reason = ""

    for fund in funds:
        print(c(CYAN, f"── {fund} {'(dry-run) ' if args.dry_run else ''}──────────────────────────"))

        # Load distribution JSON
        json_path = os.path.join(dist_dir, f"{fund}.json")
        if not os.path.exists(json_path):
            print(c(YELLOW, f"  WARNING: {fund}.json not found in {dist_dir} — skipping."))
            print(c(DIM,    f"           Run Step 1 first."))
            continue

        with open(json_path) as f:
            dist_data = json.load(f)

        # Extract ROC rows for this year
        roc_rows = build_roc_rows(dist_data, year)
        if not roc_rows:
            print(c(DIM,    f"  No ROC rows found for {year} in {fund}.json"))
            continue

        print(c(CYAN,   f"  Found {len(roc_rows)} ROC row(s) in distribution JSON for {year}"))

        # Find sheet
        if fund not in wb.sheetnames:
            print(c(YELLOW, f"  WARNING: Sheet '{fund}' not found in {os.path.basename(acb_path)} — skipping."))
            print(c(DIM,    f"           Sheet names found: {', '.join(wb.sheetnames)}"))
            continue

        ws = wb[fund]
        last_data_row  = find_last_data_row(ws)
        existing_roc   = get_sheet_roc_rows(ws, last_data_row)

        print(c(DIM,    f"  Sheet has {last_data_row - 1} data rows, {len(existing_roc)} existing ROC entries"))

        # Insert
        ins, skip, warn, errs = insert_roc_rows(
            ws, roc_rows, existing_roc, year, args.dry_run, fund
        )

        # Formula integrity check (skip on dry-run)
        if not args.dry_run and ins > 0:
            new_last = find_last_data_row(ws)
            integrity_errors = check_formula_integrity(ws, 2, new_last)
            if integrity_errors:
                print(c(RED,   f"\n  INTEGRITY CHECK FAILED for {fund}:"))
                for e in integrity_errors:
                    print(c(RED,   f"    {e}"))
                rollback        = True
                rollback_reason = f"Formula integrity check failed on sheet {fund}"
                break
            else:
                print(c(GREEN, f"  Formula integrity check passed ✓"))

        total_inserted += ins
        total_skipped  += skip
        total_warned   += warn
        print()

    # ── Rollback or save ──────────────────────────────────────────────────────
    if rollback:
        print(c(RED,   f"\n  ROLLING BACK — {rollback_reason}"))
        if backup_path and os.path.exists(backup_path):
            shutil.copy2(backup_path, acb_path)
            print(c(DIM,   f"  Restored from backup: {os.path.basename(backup_path)}"))
        sys.exit(1)

    if not args.dry_run and total_inserted > 0:
        wb.save(acb_path)
        print(c(GREEN, f"  Saved: {os.path.basename(acb_path)}"))

    # ── Summary ───────────────────────────────────────────────────────────────
    print()
    print(c(BOLD + YELLOW, "── Summary " + "─" * 50))
    print(c(GREEN,  f"  Inserted : {total_inserted}"))
    print(c(DIM,    f"  Skipped  : {total_skipped} (duplicates)"))
    if total_warned > 0:
        print(c(YELLOW, f"  Warnings : {total_warned} (existing ROC with different amount — review manually)"))
    if args.dry_run:
        print()
        print(c(YELLOW,  "  Dry-run complete — no changes were made."))
    elif total_inserted == 0 and total_warned == 0:
        print()
        print(c(DIM,     "  Nothing to do — all ROC rows already present."))
    else:
        print()
        if backup_path:
            print(c(DIM, f"  Backup   : {os.path.basename(backup_path)}"))

    if total_warned > 0:
        print()
        print(c(YELLOW,  "  Done. Open the spreadsheet and verify the inserted rows look correct."))
        print(c(YELLOW,  "  If anything looks wrong, restore the backup by renaming it back to"))
        print(c(YELLOW, f"  {os.path.basename(acb_path)}"))
        sys.exit(2)  # warnings present — run_t3.py will show "completed with warnings"

if __name__ == "__main__":
    main()

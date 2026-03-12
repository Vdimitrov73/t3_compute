"""
build_assets.py
---------------
Reads the ACB spreadsheet (XLS/XLSX) and the T3 JSON files already generated
by parse_t3_pdfs.py, then produces one JSON per brokerage account showing the
share balance on each distribution record date.

Requirements:
    pip install xlrd          # for .xls files  (pip install xlrd==1.2.0)
    pip install openpyxl      # for .xlsx files  (already installed)

Usage:
    python build_assets.py [--config config.json] [--year 2025] [--funds VBAL ZCN]

Account periods are loaded from account_periods.json (same directory as config).
If a fund is not found in account_periods.json, the script falls back to
interactive prompts.

account_periods.json format:
{
  "2025": {
    "VBAL": [
      {"start": "2021-12-29", "account": "TDDI"},
      {"start": "2025-10-01", "account": "Wealthsimple Cash"}
    ],
    ...
  }
}
"""

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, date

# ── Config ────────────────────────────────────────────────────────────────────

def load_config(path="config.json"):
    with open(path) as f:
        cfg = json.load(f)
    _set_derived_paths(cfg, cfg["tax_year"])
    return cfg

def _set_derived_paths(cfg, year):
    cfg["pdf_dir"]    = os.path.join(cfg["base_dir"], year)
    cfg["dist_dir"]   = os.path.join(cfg["base_dir"], year, "distributions")
    cfg["assets_dir"] = os.path.join(cfg["base_dir"], year, "assets")
    cfg["output_txt"] = os.path.join(cfg["base_dir"], year, f"T3_results_{year}.txt")

def parse_args():
    parser = argparse.ArgumentParser(description="Build per-account share-balance JSONs from ACB spreadsheet.")
    parser.add_argument("--config", default="config.json", help="Path to config.json")
    parser.add_argument("--year",   help="Override tax year in config")
    parser.add_argument("--funds",  nargs="+", help="Override fund list (e.g. --funds VBAL ZCN)")
    return parser.parse_args()

def load_account_periods(config_path):
    """Load account_periods.json from same directory as config.json. Returns {} if not found."""
    periods_path = os.path.join(os.path.dirname(os.path.abspath(config_path)), "account_periods.json")
    if os.path.exists(periods_path):
        with open(periods_path) as f:
            return json.load(f)
    return {}

# ── Date / number helpers ─────────────────────────────────────────────────────

DATE_FORMATS = ["%Y-%b-%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]

def parse_date(val):
    """Try to parse a cell value as a date. Returns date or None."""
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, float):
        try:
            import xlrd
            # datemode 0 = Windows 1900-epoch (correct for all modern .xls files)
            # .xlsx files never produce float dates here (openpyxl returns date objects)
            t = xlrd.xldate_as_tuple(val, 0)
            return date(t[0], t[1], t[2])
        except Exception:
            pass
    if isinstance(val, str):
        val = val.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
    return None

def parse_number(val):
    """Return float from a cell value, or None."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace("$", "").replace(",", "")
        try:
            return float(val)
        except ValueError:
            pass
    return None

# ── Spreadsheet loader ────────────────────────────────────────────────────────

def load_sheet(xls_path, sheet_name):
    """
    Load a named sheet from an XLS or XLSX file.
    Returns list of rows (each row is a list of raw cell values), or None if sheet not found.

    NOTE: opened WITHOUT data_only so formula strings are readable. Share balance is
    computed by build_share_balance_map() from Buy/Sell rows instead of reading col G,
    because newly inserted rows (written by update_acb.py) have uncached formula strings
    that data_only=True would return as None.
    """
    ext = os.path.splitext(xls_path)[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(xls_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        return [[cell.value for cell in row] for row in ws.iter_rows()]
    else:
        try:
            import xlrd
        except ImportError:
            print("\nERROR: xlrd is required to read .xls files.")
            print("  Run:  pip install xlrd==1.2.0")
            sys.exit(1)
        wb = xlrd.open_workbook(xls_path)
        if sheet_name not in wb.sheet_names():
            return None
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                row.append(cell.value)
            rows.append(row)
        return rows

def build_share_balance_map(rows, col_date, col_txn=1, col_shares=3):
    """
    Compute the running share balance for every row by walking Buy/Sell entries.
    Returns a list of (date, balance) pairs in row order, skipping non-date rows.

    This avoids reading col G (Share Balance) which is a formula column — openpyxl
    returns None for formula cells in files that have never been opened by Excel
    (e.g. rows freshly inserted by update_acb.py).
    """
    balance = 0.0
    result  = []  # list of (date, balance_after_this_row)
    for row in rows:
        if len(row) <= max(col_date, col_txn, col_shares):
            continue
        d = parse_date(row[col_date])
        if d is None:
            continue
        txn = str(row[col_txn]).strip() if row[col_txn] is not None else ""
        shares_val = parse_number(row[col_shares]) or 0.0
        if txn == "Buy":
            balance = round(balance + shares_val, 4)
        elif txn == "Sell":
            balance = round(balance - shares_val, 4)
        # ROC rows: balance unchanged
        result.append((d, balance))
    return result

def get_share_balance_on_date(rows, target_date, col_date, col_share, fund=""):
    """
    Return the share balance for the last row on or before target_date.

    Always computes from Buy/Sell transactions (col B/D) — works correctly
    whether or not the file has been opened in Excel after Step 2 insertions.

    If col G has a cached value for the same row, cross-checks against it.
    A mismatch is a warning (possible spreadsheet integrity issue), not an error.
    """
    balance_map = build_share_balance_map(rows, col_date)
    best_date   = None
    best_share  = None
    best_row_idx = None

    for i, (d, bal) in enumerate(balance_map):
        if d <= target_date:
            if best_date is None or d >= best_date:
                best_date    = d
                best_share   = bal
                best_row_idx = i

    # Cross-check against cached col G value if available
    if best_share is not None and best_row_idx is not None:
        # Find the actual spreadsheet row that corresponds to best_row_idx
        data_rows = [row for row in rows if len(row) > col_date and parse_date(row[col_date]) is not None]
        if best_row_idx < len(data_rows):
            cached = parse_number(data_rows[best_row_idx][col_share]) if len(data_rows[best_row_idx]) > col_share else None
            if cached is not None and abs(cached - best_share) > 0.001:
                label = f"[{fund}] " if fund else ""
                print(f"  {label}WARNING: Share balance mismatch on {best_date} — "
                      f"computed {best_share:,.4f} from Buy/Sell vs cached {cached:,.4f} in col G. "
                      f"Check your spreadsheet for missing or incorrect transactions.")
                return best_share, True   # (value, had_warning)

    return best_share, False   # (value, had_warning)

# ── Account period helpers ────────────────────────────────────────────────────

def get_account_periods(fund, tax_year, all_periods, record_dates):
    """
    Return list of (start_date, account_name) for a fund.
    Tries account_periods.json first; falls back to interactive prompts if not found.
    """
    year_periods = all_periods.get(tax_year, {})
    if fund in year_periods:
        periods = [(datetime.strptime(p["start"], "%Y-%m-%d").date(), p["account"])
                   for p in year_periods[fund]]
        print(f"  Loaded {len(periods)} period(s) from account_periods.json")
        return sorted(periods, key=lambda x: x[0])

    # Fall back to interactive
    print(f"  '{fund}' not found in account_periods.json — switching to interactive input.")
    return prompt_account_periods(fund, record_dates)

def prompt_account_periods(fund, record_dates):
    """
    Interactively prompt for account periods.
    Returns sorted list of (start_date, account_name), or None if user skips.
    """
    print(f"\n  Fund: {fund}  (record dates: {', '.join(str(d) for d in record_dates)})")
    periods  = []
    period_num = 1
    while True:
        prompt = (f"    Period {period_num} start date (YYYY-MM-DD)"
                  + (" — or press Enter to skip this fund" if period_num == 1 else " — or press Enter when done")
                  + ": ")
        raw = input(prompt).strip()
        if not raw:
            if period_num == 1:
                return None
            break
        try:
            start = datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            print("    Invalid date format, try again.")
            continue
        account = input(f"    Period {period_num} account name: ").strip()
        if not account:
            print("    Account name cannot be empty, try again.")
            continue
        periods.append((start, account))
        period_num += 1

    return sorted(periods, key=lambda x: x[0])

def account_for_date(periods, d):
    """Return the account name active on date d."""
    active = None
    for start, name in periods:
        if d >= start:
            active = name
    return active

def sanitize_filename(name):
    """Make a string safe for use as a filename."""
    return re.sub(r'[\\/:*?"<>|]', "_", name)

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()
    cfg  = load_config(args.config)

    # Apply CLI overrides
    if args.year:
        cfg["tax_year"] = args.year
        _set_derived_paths(cfg, args.year)

    tax_year   = cfg["tax_year"]
    dist_dir   = cfg["dist_dir"]
    assets_dir = cfg["assets_dir"]
    xls_path   = cfg["acb_spreadsheet"]
    col_date   = cfg["col_indices"]["date"]
    col_share  = cfg["col_indices"]["share_balance"]

    os.makedirs(assets_dir, exist_ok=True)

    # ── Prerequisite check ────────────────────────────────────────────────────
    if not os.path.exists(xls_path):
        print(f"ERROR: ACB spreadsheet not found: {xls_path}")
        print("       Update 'acb_spreadsheet' in config.json.")
        sys.exit(1)
    if not os.path.isdir(dist_dir):
        print(f"ERROR: Distributions folder not found: {dist_dir}")
        print("       Run Step 1 (parse_t3_pdfs.py) first.")
        sys.exit(1)
    dist_jsons = [f for f in os.listdir(dist_dir) if re.match(r'^[A-Z]+\.json$', f)]
    if not dist_jsons:
        print(f"ERROR: No distribution JSONs found in: {dist_dir}")
        print("       Run Step 1 (parse_t3_pdfs.py) first.")
        sys.exit(1)

    # Load account periods from account_periods.json
    all_periods = load_account_periods(args.config)

    # Discover funds: use CLI override, else config list, else scan dist_dir for JSONs
    if args.funds:
        funds = args.funds
    elif cfg.get("funds"):
        funds = cfg["funds"]
    else:
        json_files = [f for f in os.listdir(dist_dir) if re.match(r'^[A-Z]+\.json$', f)]
        if not json_files:
            print(f"No T3 JSON files found in {dist_dir}")
            print("Run parse_t3_pdfs.py first.")
            sys.exit(1)
        funds = sorted(f.replace(".json", "") for f in json_files)

    print(f"Tax year       : {tax_year}")
    print(f"Dist dir       : {dist_dir}")
    print(f"Assets dir     : {assets_dir}")
    print(f"ACB spreadsheet: {xls_path}")
    print(f"Funds          : {', '.join(funds)}")
    print()

    fund_data    = {}
    all_accounts = []
    warning_count = 0

    for fund in funds:
        json_path = os.path.join(dist_dir, f"{fund}.json")
        if not os.path.exists(json_path):
            print(f"  WARNING: {json_path} not found — skipping {fund}.")
            continue

        with open(json_path) as f:
            t3 = json.load(f)

        record_dates = []
        for dist in t3["distributions"]:
            try:
                record_dates.append(datetime.strptime(dist["recordDate"], "%Y-%m-%d").date())
            except Exception:
                pass

        if not record_dates:
            print(f"  {fund}: no distributions, skipping.")
            continue

        print(f"Processing {fund}  ({len(record_dates)} record date(s))...")

        # Get account periods (from JSON or interactive)
        periods = get_account_periods(fund, tax_year, all_periods, record_dates)
        if periods is None:
            print(f"  Skipping {fund}.")
            continue

        # Track account names in order of first appearance
        for _, acct in periods:
            if acct not in all_accounts:
                all_accounts.append(acct)

        # Load ACB sheet
        rows = load_sheet(xls_path, fund)
        if rows is None:
            print(f"  WARNING: Sheet '{fund}' not found in {xls_path} — skipping.")
            print(f"           Sheet name must match the fund ticker exactly.")
            warning_count += 1
            continue

        # Build date → shares map
        shares_by_date = {}
        for d in record_dates:
            shares, had_mismatch = get_share_balance_on_date(rows, d, col_date, col_share, fund=fund)
            if had_mismatch:
                warning_count += 1
            if shares is None:
                print(f"  WARNING: Could not find share balance for {fund} on {d}")
                print(f"           Make sure a Buy/Sell row exists on or before this date.")
                shares = 0.0
                warning_count += 1
            shares_by_date[d] = round(shares, 4)

        fund_data[fund] = {
            "periods": periods,
            "dates":   record_dates,
            "shares":  shares_by_date,
        }

    if not fund_data or not all_accounts:
        print("\nNo data collected. Exiting.")
        if warning_count > 0:
            sys.exit(2)
        return

    # Build per-account data: every account gets every distribution date for every fund.
    # Dates outside an account's active period get ownedShares = 0.
    accounts_data = defaultdict(lambda: defaultdict(list))
    for fund, fd in fund_data.items():
        for d in fd["dates"]:
            active_account = account_for_date(fd["periods"], d)
            for acct in all_accounts:
                shares = fd["shares"][d] if acct == active_account else 0.0
                accounts_data[acct][fund].append({
                    "date":        str(d),
                    "ownedShares": shares
                })

    # Write one JSON per account
    print()
    for account, funds_data in accounts_data.items():
        output = {"account": account}
        for fund, entries in sorted(funds_data.items()):
            output[fund] = entries

        filename = f"{sanitize_filename(account)}_{tax_year}.json"
        out_path = os.path.join(assets_dir, f"{sanitize_filename(account)}_{tax_year}.json")

        with open(out_path, "w") as f:
            lines = ['{\n']
            items = list(output.items())
            for i, (key, val) in enumerate(items):
                comma = "," if i < len(items) - 1 else ""
                if key == "account":
                    lines.append(f'  "account": "{val}",\n')
                else:
                    lines.append(f'  "{key}": [\n')
                    for j, entry in enumerate(val):
                        entry_comma = "," if j < len(val) - 1 else ""
                        lines.append(
                            f'    {{ "date": "{entry["date"]}", '
                            f'"ownedShares": {entry["ownedShares"]} }}{entry_comma}\n'
                        )
                    lines.append(f'  ]{comma}\n')
            lines.append('}')
            f.writelines(lines)

        print(f"Saved: {out_path}")

    if warning_count > 0:
        print(f"\n  WARNING: {warning_count} warning(s) — share balances may be incomplete.")
        print(f"  Review the warnings above before running Step 4.")
        sys.exit(2)

    print("\nDone.")


if __name__ == "__main__":
    main()
